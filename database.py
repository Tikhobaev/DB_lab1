import os
import zipfile

import openpyxl


class FilmsDatabase:
    def __init__(self, filename, empty=False):
        if empty:
            self.filename = filename
            self.db = []
            self.id_dicts = {}
            self.name = ''
        else:
            if os.path.exists(filename):
                self.filename = filename
                rows = []
                with open(filename, 'r') as data_file:
                    for line in data_file:
                        raw_row = line.split(';')
                        rows.append([raw_row[0].replace('^', ''), int(raw_row[1]), float(raw_row[2]), int(raw_row[3])])
                if rows:
                    self.db = rows
                    self.id_dicts = dict([[rows[i][3], i] for i in range(0, len(rows))])
                else:
                    self.db = []
                    self.id_dicts = {}
            else:
                self.name = filename
                self.db = []
                self.id_dicts = {}
                if self.name.endswith(r'.hdb'):
                    self.filename = self.name
                else:
                    self.filename = f'{self.name}.hdb'

        self.node_size_in_bytes = 48
        self.field_lengths = [30, 4, 3, 7]

    def change(self, new_db):
        self.filename = new_db.filename
        self.db = new_db.db
        self.id_dicts = new_db.id_dicts

    def save_db(self, filename):
        if filename:
            with open(filename, 'w') as data_file:
                counter = 0
                for node in self.db:
                    node[0] = node[0] + (self.field_lengths[0] - len(node[0])) * '^'
                    node[1] = str(node[1])
                    node[2] = (self.field_lengths[2] - len(str(node[2]))) * '0' + str(node[2])
                    node[3] = str(node[3])
                    counter += 1
                    data_file.write(';'.join(node))
                    if counter < len(self.db):
                        data_file.write('\n')

    @staticmethod
    def create_db(filename):
        with open(filename, 'w') as f:
            pass
        return True

    @staticmethod
    def open_db(filename):
        if os.path.exists(filename):
            return FilmsDatabase(filename)
        return None

    @staticmethod
    def del_db(filename):
        if os.path.exists(filename):
            os.remove(filename)
            return True
        return None

    def make_backup(self, filename):
        if self.filename:
            newzip = zipfile.ZipFile(filename, 'w')
            newzip.write(self.filename)
            newzip.close()

    @staticmethod
    def from_backup(filename_zip, final_dir):
        arch = zipfile.ZipFile(filename_zip, 'r')
        filename = filename_zip.split('/')[-1].replace('.zip', '')
        arch.extract(filename, 'extracted')
        arch.close()
        os.replace(f'extracted/{filename}', final_dir + filename)
        os.rmdir(f'extracted')

    def save_to_xlsx(self, filename):
        wb = openpyxl.Workbook()
        wb_name = filename.split('/')[-1].replace('.xlsx', '')
        wb.create_sheet(title=wb_name, index=0)
        sheet = wb[wb_name]
        for i in range(0, len(self.db)):
            for j in range(0, 4):
                value = self.db[i][j]
                cell = sheet.cell(row=i+1, column=j+1)
                cell.value = value
        wb.save(filename)

    def search_by_name(self, name) -> list:
        result = []
        for row in self.db:
            if row[0] == name:
                result.append(row)
        return result

    def search_by_year(self, year) -> list:
        year = int(year)
        result = []
        for row in self.db:
            if row[1] == year:
                result.append(row)
        return result

    def search_by_rate(self, rate) -> list:
        rate = float(rate)
        result = []
        for row in self.db:
            if row[2] == rate:
                result.append(row)
        return result

    def search_by_id(self, node_id) -> list:
        node_id = int(node_id)
        result = []
        if self.id_dicts.get(node_id) is not None:
            result.append(self.db[self.id_dicts[node_id]])
        return result

    def add(self, new_node: list):
        # if id value already exists
        if self.id_dicts.get(int(new_node[3])) is not None:
            return
        try:
            with open(self.filename, 'a') as data_file:
                self.db.append([new_node[0], int(new_node[1]), float(new_node[2]), int(new_node[3])])
                self.id_dicts[new_node[3]] = len(self.db) - 1
                if len(new_node[0]) < self.field_lengths[0]:
                    new_node[0] = new_node[0] + '^' * (self.field_lengths[0] - len(new_node[0]))
                if len(new_node[2]) != self.field_lengths[2]:
                    new_node[2] = '0' * (self.field_lengths[2] - len(new_node[2])) + new_node[2]
                if self.db:
                    data_to_insert = '\n' + ';'.join(new_node)
                    if len(data_to_insert) == self.node_size_in_bytes:
                        data_file.write(data_to_insert)
                else:
                    data_to_insert = ';'.join(new_node)
                    if len(data_to_insert) == self.node_size_in_bytes - 1:
                        data_file.write(data_to_insert)
        except Exception:
            pass

    def delete_node(self, node_id):
        node_id = int(node_id)
        if self.id_dicts.get(node_id) is None:
            print('This id is not exists')
            return
        with open(self.filename, 'r+') as data_file:
            data_file.seek(self.node_size_in_bytes * (len(self.db) - 1))
            final_node = data_file.readline().split(';')
            final_node_id = int(final_node[3])
            data_file.seek(self.node_size_in_bytes * self.id_dicts[node_id])
            data_file.write(';'.join(final_node))
            if len(self.db) > 1:
                data_file.truncate(self.node_size_in_bytes * (len(self.db) - 1) - 1)
            else:
                data_file.truncate(self.node_size_in_bytes * (len(self.db) - 1))
        self.db[self.id_dicts[node_id]] = self.db[self.id_dicts[final_node_id]]
        self.db.pop()
        self.id_dicts[final_node_id] = self.id_dicts[node_id]
        del self.id_dicts[node_id]
        pass

    def change_node(self, updated_data: list):
        """

        :param updated_data: list which consist of 4 elements - updated values of fields for the node
        :return:
        """
        node_id = int(updated_data[3])
        if self.id_dicts.get(node_id) is None:
            return False
        self.db[self.id_dicts[node_id]] = [updated_data[0], int(updated_data[1]), float(updated_data[2]), int(updated_data[3])]
        if len(updated_data[0]) < self.field_lengths[0]:
            updated_data[0] += '^' * (self.field_lengths[0] - len(updated_data[0]))
        if len(updated_data[2]) != self.field_lengths[2]:
            updated_data[2] = '0' * (self.field_lengths[2] - len(updated_data[2])) + updated_data[2]
        with open(self.filename, 'r+') as data_file:
            data_file.seek(self.node_size_in_bytes * self.id_dicts[node_id])
            data_file.write(';'.join(updated_data))
        return True

